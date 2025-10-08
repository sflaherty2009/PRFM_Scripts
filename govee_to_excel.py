#!/usr/bin/env python3
import os
import sys
import json
import datetime as dt
from zoneinfo import ZoneInfo
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ---- Config ----
API_KEY = os.getenv("GOVEE_API_KEY")  # set this in your shell; don't hardcode
if not API_KEY:
    print("ERROR: Set GOVEE_API_KEY in your environment.")
    sys.exit(1)

# Your devices (from your output)
DEVICES = [
    {
        "name": "Lab Fridge 1",
        "sku":  "H5111",
        "id":   "FE:32:D0:03:81:C6:31:81",
    },
    {
        "name": "Lab Fridge 2",
        "sku":  "H5111",
        "id":   "09:EC:D0:03:80:86:66:33",
    },
]

# Units: Govee returns temps that (for your fridges) look like °F (e.g., 33.8).
# If yours are actually °C, flip DEFAULT_UNIT to "C".
DEFAULT_UNIT = os.getenv("GOVEE_TEMP_UNIT", "F").upper()  # "F" or "C"

# Output Excel
XLSX_PATH = "govee_temps.xlsx"
SHEET_NAME = "readings"

# ---- Helpers ----
def c_to_f(c): return (c * 9/5) + 32
def f_to_c(f): return (f - 32) * 5/9

def ensure_workbook(path):
    try:
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(["timestamp", "device_name", "device_id", "sku", "temp_f", "temp_c"])
        else:
            ws = wb[SHEET_NAME]
        return wb, ws
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["timestamp", "device_name", "device_id", "sku", "temp_f", "temp_c"])
        return wb, ws

def autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def fetch_temp_f_and_c(session, sku, device_id):
    url = "https://openapi.api.govee.com/router/api/v1/device/state"
    headers = {"Content-Type": "application/json", "Govee-API-Key": API_KEY}
    body = {
        "requestId": "uuid",
        "payload": {"sku": sku, "device": device_id}
    }
    r = session.post(url, headers=headers, json=body, timeout=15)
    r.raise_for_status()
    data = r.json()

    # Find the sensorTemperature capability
    caps = data.get("payload", {}).get("capabilities", [])
    val = None
    for cap in caps:
        if cap.get("instance") == "sensorTemperature":
            val = cap.get("state", {}).get("value")
            break
    if val is None:
        raise RuntimeError(f"No temperature value found for {device_id}")

    # Normalize to both F and C
    if DEFAULT_UNIT == "F":
        temp_f = float(val)
        temp_c = round(f_to_c(temp_f), 2)
    else:
        temp_c = float(val)
        temp_f = round(c_to_f(temp_c), 2)

    return temp_f, temp_c

def main():
    tz = ZoneInfo("America/Chicago")
    timestamp = dt.datetime.now(tz).isoformat(timespec="seconds")

    wb, ws = ensure_workbook(XLSX_PATH)

    with requests.Session() as s:
        for d in DEVICES:
            temp_f, temp_c = fetch_temp_f_and_c(s, d["sku"], d["id"])
            ws.append([timestamp, d["name"], d["id"], d["sku"], temp_f, temp_c])

    autosize(ws)
    wb.save(XLSX_PATH)
    print(f"Appended {len(DEVICES)} rows to {XLSX_PATH} at {timestamp}")

if __name__ == "__main__":
    main()
